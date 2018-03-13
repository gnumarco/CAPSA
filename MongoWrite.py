from pymongo import MongoClient
import openpyxl

# connection
uri = "mongodb://Dayana:capsa@172.31.30.94:27017/?authSource=admin"
client = MongoClient(uri)

db = client.get_database("DATASCIENCE")
MyDoc=[]

#mode = 1  # Eroski
#mode = 2 #ECI
#mode = 3  # CRF
mode = 4 # MYM

if mode==1:
    doc=openpyxl.load_workbook('C:\\Users\\tr5568\\Desktop\\DAYANA\\CAPSA\\Softwares\\PROMOCIONES_EROSKI_LYB_DDLL_2016_1802_VersII.xlsx')
    hoja=doc.get_sheet_by_name("Promociones")

if mode==2:
    doc=openpyxl.load_workbook('C:\\Users\\tr5568\\Desktop\\DAYANA\\CAPSA\\Softwares\\PROMOCIONES_ECI_2015_1710_VersIII.xlsx')
    hoja=doc.get_sheet_by_name("Formato")

if mode == 3:
    doc = openpyxl.load_workbook(
        'C:\\Users\\tr5568\\Desktop\\DAYANA\\CAPSA\\Softwares\\PROMOCIONES_CRF_LYB_DDLL_2015_1709.xlsx')
    hoja = doc.get_sheet_by_name("PROMOCIONES_CRF")

if mode == 4:
    doc = openpyxl.load_workbook(
        'C:\\Users\\tr5568\\Desktop\\DAYANA\\CAPSA\\Softwares\\PROMOCIONES_HLR.xlsx')
    hoja = doc.get_sheet_by_name("PROMOCIONES_HLR")


i=0
for row in hoja.rows:

    if(i>0):
        MyDoc.append({"Unique Promo Code": row[0].value, "Client Code": row[1].value, "Enseña Code":row[2].value, "Initial Date": row[3].value, "Final Date": row[4].value, "Brand Code": row[5].value, "Family Code": row[6].value, "FAMAPO Code": row[7].value, "Action Abbreviation": row[8].value,
                      "Animation 1": row[9].value, "Animation 2": row[10].value, "Animation 3": row[11].value, "Base PVP": row[12].value, "Grouped PVP": row[13].value, "Thematic": row[14].value, "Location": row[15].value, "Highlighted": row[16].value, "Marketing": row[17].value, "Mode":mode})

    i+=1

print(MyDoc)

write_result = db.Promos.insert_many(MyDoc)